from django.forms import modelformset_factory
from django.utils.translation import gettext_lazy as _

import xlsxwriter
from openpyxl import load_workbook
from io import BytesIO

from .views_common import *
from .FieldMap import standard_field_map
from .import_utils import *

class CommonForm(forms.Form):
    license_check_note = forms.CharField(
		label=_('Common License Check Note'),
		help_text=_('(shown on License Check page for all Categories)'),
		max_length=240,
		required=False,
		widget=forms.Textarea(attrs={'rows':4, 'cols':80}))

CompetitionCategoryOptionFormSet = modelformset_factory(
	CompetitionCategoryOption, fields='__all__',
	widgets={
		'note': forms.TextInput(attrs={'size':40}),
		'competition': forms.HiddenInput(),
		'category': forms.HiddenInput(),
	},
	extra=0,
)
	
@access_validation()
@user_passes_test( lambda u: u.is_superuser )
def SetLicenseChecks( request, competitionId ):
	competition = get_object_or_404( Competition, pk=competitionId )
	CompetitionCategoryOption.normalize( competition )
	
	ccos_query = competition.competitioncategoryoption_set.all().order_by('category__sequence').select_related('category')
	
	if request.method == 'POST':
		if 'set-all-submit' in request.POST:
			ccos_query.update( license_check_required=True )
			return HttpResponseRedirect( '.' )
			
		if 'clear-all-submit' in request.POST:
			ccos_query.update( license_check_required=False )
			return HttpResponseRedirect( '.' )
			
		if 'import-excel-submit' in request.POST:
			return HttpResponseRedirect( pushUrl(request, 'UploadCCOs', competition.id) )
		
		form_common = CommonForm( request.POST, prefix='common' )
		form_set = CompetitionCategoryOptionFormSet( request.POST, prefix='cco' )
		if form_set.is_valid() and form_common.is_valid():
			competition.license_check_note = form_common.cleaned_data['license_check_note']
			competition.save()
			form_set.save()
			
			if 'export-excel-submit' in request.POST:
				xl = ccos_to_excel( competition )
				response = HttpResponse(xl, content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
				response['Content-Disposition'] = 'attachment; filename={}-CCO-{}.xlsx'.format(
					timezone.now().strftime('%Y-%m-%d'),
					utils.cleanFileName(competition.name),
				)
				return response

			return HttpResponseRedirect( getContext(request, 'cancelUrl') )
	else:
		form_common = CommonForm( initial={'license_check_note':competition.license_check_note}, prefix='common' )
		form_set = CompetitionCategoryOptionFormSet( queryset=ccos_query, prefix='cco' )

	return render( request, 'cco_license_check_form.html', locals() )

def ccos_to_excel( competition ):
	CompetitionCategoryOption.normalize( competition )
	ccos_query = competition.competitioncategoryoption_set.all().order_by('category__sequence').select_related('category')
		
	output = BytesIO()
	wb = xlsxwriter.Workbook( output, {'in_memory': True} )
	title_format = wb.add_format( dict(bold=True) )
	
	sheet_name = 'RaceDB-CCO'
	ws = wb.add_worksheet(sheet_name)
	
	row = 0
	
	ws.write( row, 0, '{}'.format(_('Category')), title_format )
	ws.write( row, 1, '{}'.format(_('License Check Required')), title_format )
	ws.write( row, 2, '{}'.format(_('Note')), title_format )
	
	for cco in ccos_query:
		row += 1
		ws.write( row, 0, cco.category.code_gender )
		ws.write( row, 1, cco.license_check_required )
		ws.write( row, 2, cco.note )
			
	sheet_name = 'RaceDB-Common'
	ws = wb.add_worksheet(sheet_name)
	row = 0
	
	ws.write( row, 0, '{}'.format(_('License Check Note')), title_format )
	row += 1
	ws.write( row, 0, competition.license_check_note )
	
	wb.close()
	return output.getvalue()

def ccos_from_excel( competition, worksheet_contents, sheet_name=None ):
	CompetitionCategoryOption.normalize( competition )
	ccos_query = competition.competitioncategoryoption_set.all().order_by('category__sequence').select_related('category')

	wb = load_workbook( filename=BytesIO(worksheet_contents), read_only=True, data_only=True )
	
	try:
		sheet_name = sheet_name or wb.sheetnames[0]
		ws = wb[sheet_name]
	except Exception:
		return
	
	ccos = {cco.category.code_gender.lower():cco for cco in ccos_query}
	
	ifm = standard_field_map()

	for r, row in enumerate(ws.iter_rows()):
		if r == 0:
			# Get the header fields from the first row.
			fields = ['{}'.format(v.value).strip() for v in row]
			ifm.set_headers( fields )
			continue
		
		values = [c.value for c in row]
		v = ifm.finder( values )
		code_gender = to_str(v('category_code', None))
		if not code_gender:
			continue
		try:
			cco = ccos[code_gender.strip().lower()]
		except (ValueError, KeyError):
			continue
		
		license_check_required = to_bool(v('license_check_required', None))		
		if license_check_required is not None:
			cco.license_check_required = license_check_required
		
		note = v('note', None)
		if note is not None:
			cco.note = '{}'.format(note)
		
		cco.save()
	
	sheet_name = 'RaceDB-Common'
	try:
		ws = wb[sheet_name]
	except Exception:
		return
	
	ifm = standard_field_map()

	for r, row in enumerate(ws.iter_rows()):
		if r == 0:
			# Get the header fields from the first row.
			fields = ['{}'.format(v.value).strip() for v in row]
			ifm.set_headers( fields )
			continue
		
		values = [c.value for c in row]
		v = ifm.finder( values )
		
		license_check_note = to_str(v('license_check_note', None))
		if license_check_note is not None:
			competition.license_check_note = license_check_note
			
		competition.save()

#-----------------------------------------------------------------------
@autostrip
class UploadCCOForm( Form ):
	excel_file = forms.FileField( required=True, label=_('Excel Spreadsheet (*.xlsx)') )
	
	def __init__( self, *args, **kwargs ):
		super().__init__( *args, **kwargs )
		self.helper = FormHelper( self )
		self.helper.form_action = '.'
		self.helper.form_class = 'form-inline'
		
		self.helper.layout = Layout(
			Row(
				Col( Field('excel_file', accept=".xls,.xlsx"), 8),
			),
		)
		addFormButtons( self, OK_BUTTON | CANCEL_BUTTON, cancel_alias=_('Done') )

def handle_upload( competition, excel_contents ):
	worksheet_contents = excel_contents.read()
	ccos_from_excel( competition, worksheet_contents=worksheet_contents, )

@access_validation()
@user_passes_test( lambda u: u.is_superuser )
def UploadCCOs( request, competitionId ):
	competition = get_object_or_404( Competition, pk=competitionId )
	
	if request.method == 'POST':
		form = UploadCCOForm(request.POST, request.FILES)
		if form.is_valid():
			handle_upload( competition, request.FILES['excel_file'] )
			return HttpResponseRedirect(getContext(request,'cancelUrl'))
	else:
		form = UploadCCOForm()
	
	return render( request, 'generic_form.html', locals() )
