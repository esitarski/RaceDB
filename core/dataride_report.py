import io
import re
import operator

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side

from django.utils.translation import gettext_lazy as _

from .views_common import *

class DatarideReportForm( Form ):
	bibs = forms.CharField( label=_('Bibs'), help_text=_('Bib finish order (paste from timing system)'), widget=forms.Textarea(attrs={'rows': 4, 'cols': 60}))
	finish_times = forms.CharField( label=_('Finish Times'), required=False, help_text=_('Finish Times (paste from timing system).  Include Times, Laps down (eg. -1, -2) and Status (eg. DNF, DNS, DSQ)'), widget=forms.Textarea(attrs={'rows': 4, 'cols': 60}) )
	
	def __init__( self, *args, **kwargs ):
		super().__init__(*args, **kwargs)
		
		self.helper = FormHelper( self )
		self.helper.form_action = '.'
		self.helper.form_class = 'form-inline'
		
		self.helper.layout = Layout(
			Row( Field('bibs') ),
			Row( Field('finish_times') ),
		)

		self.additional_buttons = [
			( 'update-submit', _('Update'), 'btn btn-success' ),
			( 'excel-submit', format_lazy( '{} {}', "🡇", _('Excel')), 'btn btn-primary' ),
		]
		addFormButtons( self, CANCEL_BUTTON, self.additional_buttons )

def get_report( wave, bibs, finish_times ):
	bibs = [int(bib) if bib.isdigit() else -1 for bib in bibs]
	if len(finish_times) < len(bibs):
		finish_times += [None] * (len(bibs) - len(finish_times))
	
	bib_participant = { p.bib:p for p in wave.get_participants_unsorted().iterator() }
	uci_ids = { p.license_holder.uci_id for p in bib_participant.values() }
	uci_lookup = {
		uci.uci_id:uci for uci in UCIRank.objects.filter(competition=wave.event.competition) if uci.uci_id in uci_ids
	}
	
	def get_participant_entry( bib, finish_time, p, rank, default_status='DNF' ):
		if finish_time:
			m = re.match( r'(\d+)h\s*(\d+)[′\']\s*(\d+)[″"]', finish_time )
			if m:
				hours, minutes, seconds = int(m.group(1)), int(m.group(2)), int(m.group(3))
				finish_time = f'{hours}:{minutes:02d}:{seconds:02d}'
		
		r = {
			'rank':			rank,
			'bib':			bib,
			'uci_id':		p.license_holder.uci_id if p else '',
			'result':		finish_time,
			'gender':		'MWO'[p.category.gender] if p else '',	# gender is M and W (not M and F).
			'sort_order':	rank,
		}
		try:
			# If there is a UCI Ranking, use it's data as that is what dataride expects.
			uci = uci_lookup[r['uci_id']]
			r.update( {
				'last_name':	utils.removeDiacritic( uci.last_name ).upper(),
				'first_name':	utils.removeDiacritic( uci.first_name ),
				'country':		uci.nation_name,
				'team':			uci.team_code,
			})
		except KeyError:
			# If no UCI Ranking, use the participant and license_holder information.
			r.update( {
				'last_name':	utils.removeDiacritic( p.license_holder.last_name ).upper() if p else '',
				'first_name':	utils.removeDiacritic( p.license_holder.first_name ) if p else '',
				'country':		p.license_holder.nation_code if p else '',
				'team':			'',
			})

		result = r['result']
		if not result:
			r['irm'] = default_status
		elif result.startswith('-'):
			r['irm'] = 'LAP'
		elif result.startswith('D'):	# DNF, DNS, DSQ
			r['irm'] = result
		else:
			r['irm'] = ''
		
		return r
	
	report = []
	bibs_seen = set()
	rank = 0
	for bib, finish_time in zip(bibs, finish_times):
		if bib not in bibs_seen:
			bibs_seen.add( bib )			
			rank += 1		
			report.append( get_participant_entry( bib, finish_time, bib_participant.get(bib, None), rank ) )

	if len(bibs_seen) != len(bib_participant):
		for p in sorted( bib_participant.values(), key=operator.attrgetter('bib') ):
			if p.bib not in bibs_seen:
				rank += 1
				report.append( get_participant_entry( p.bib, finish_time, p, rank, 'DNS' ) )
				report[-1]['missing'] = True
		
	return report

def sanitize_windows_filename( s ):
	s = re.sub(r'[<>:"/\\|?*\x00-\x1f]', '', s)
	s = s.strip('. ')
	return s or "unnamed"

def get_dataride_excel( wave, report ):
	wb = Workbook()
	ws = wb.worksheets[0]
	ws.title = "Results"
	
	header_fields = (
		('Rank',			'rank'),
		('BIB',				'bib'),
		('UCI ID',			'uci_id'),
		('Last Name',		'last_name'),
		('First Name',		'first_name'),
		('Country',			'country'),
		('Team',			'team'),
		('Gender',			'gender'),
		('Phase',			None),
		('Heat',			None),
		('Result',			'result'),
		('IRM',				'irm'),
		('Sort Order',		'sort_order'),
	)
	font_name = 'Calabri'
	font_regular = Font( name=font_name, size=11 )
	font_bold = Font( name=font_name, size=11, bold=True )
	font_title = Font( name=font_name, size=11, color="FFFFFF" )
	
	row = 1
	for col, (header, field) in enumerate(header_fields, 1):
		cell = ws.cell( row, col )
		cell.value = header
		cell.font = font_title
		cell.fill = PatternFill(fill_type="solid", fgColor="000000")
	
	font_regular = Font( name='Arial', size=11 )
	font_bold = Font( name='Arial', size=11, bold=True )
	center_fields = {'rank', 'bib', 'country', 'gender', 'sort_order'}
	for row, r in enumerate(report, 2):
		missing = r.get( 'missing', False )
		for col, (header, field) in enumerate(header_fields, 1):
			if field is not None:
				cell = ws.cell( row, col )
				cell.value = r[field]
				cell.font = font_bold if missing else font_regular
				if field == 'uci_id':
					cell.number_format = '@'	# Force Excel to keep this as a string.
				if field in center_fields:
					cell.alignment = Alignment(horizontal="center")
	
	# Make all columns wide enough to show the content.
	for col in ws.columns:
		max_length = max(len(str(cell.value or "")) for cell in col)
		ws.column_dimensions[col[0].column_letter].width = max_length + 4

	# Add auto-filter to all culumns.
	ws.auto_filter.ref = ws.dimensions
		
	# Make the rows alternating colors.
	
	thin = Side(style='thin', color='000000')
	border = Border(left=thin, right=thin, top=thin, bottom=thin)

	fill_even = PatternFill(fill_type="solid", fgColor="FFFFFF")  # white
	fill_odd  = PatternFill(fill_type="solid", fgColor="EEF2FF")  # light blue

	for row in ws.iter_rows(min_row=2):  # skip header
		for cell in row:
			cell.fill = fill_odd if (row[0].row & 1) else fill_even
			cell.border = border
		
	return wb

@access_validation()
@user_passes_test( lambda u: u.is_superuser )
def DatarideReport( request, waveId ):
	wave = get_object_or_404( Wave, pk=waveId )
	event = wave.event
	competition = event.competition
	
	results = list( wave.get_results() )
	if results:
		status_value = {
			cFinisher:	'',
			cDNF:		'DNF',
			cDNS:		'DNS',
			cDQ:		'DSQ',
		}
		
		bibs = []
		finish_times = []
		laps_max = None
		for r in results:
			if laps_max is None:
				if r.status != cFinisher:
					break
				laps_max = r.get_num_laps_fast()
			
			bibs.append( r.participant.bib )
			if r.get_num_laps_fast() < laps_max:
				finish_times.append( r.get_num_laps_fast() - laps_max )
			elif r.finish_time:
				finish_times.append( utils.format_time( r.finish_times.total_seconds(), high_precision=True ) )
			else:
				finish_times.append( status_value.get(r.status, 'DNF') )
				
		initial = {
			'bibs': ' '.join( str(bib) for bib in bibs ),
			'finish_times': ' '.join( ft for ft in finish_time ),
		}
	else:
		initial = {}
	
	if request.method == 'POST':
		form = DatarideReportForm( request.POST )
		if form.is_valid():
			if 'ok-submit' in request.POST:
				return HttpResponseRedirect(getContext(request,'cancelUrl'))
				
			bibs = form.cleaned_data['bibs']
			finish_times = form.cleaned_data['finish_times']

			bibs = re.sub( r'[\D]', ' ', bibs).split()
			finish_times = finish_times.split()
			
			report = get_report( wave, bibs, finish_times )
			
			if 'excel-submit' in request.POST:
				wb = get_dataride_excel( wave, report )
				buffer = io.BytesIO()
				wb.save( buffer )
				buffer.seek( 0 )
				response = HttpResponse(
					buffer.read(),
					content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
				)
				fname = sanitize_windows_filename( f'{competition.name}-{wave.name}' ) + '.xlsx'
				response["Content-Disposition"] = f'attachment; filename="{fname}"'
				return response
				
	else:
		form = DatarideReportForm( initial=initial )
		report = []
	
	return render( request, 'dataride_report.html', locals() )
	
@access_validation()
@user_passes_test( lambda u: u.is_superuser )
def DatarideReports( request, eventId ):
	event = get_object_or_404( EventMassStart, pk=eventId )
	waves = event.wave_set.all()
	return render( request, 'dataride_report_list.html', locals() )
	

