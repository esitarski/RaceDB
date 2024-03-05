import sys
import datetime
from openpyxl import load_workbook
from collections import namedtuple
from models import *
from utils import toUnicode, removeDiacritic
from django.db import transaction
from django.db.models import Q
from fix_utf8 import fix_utf8
import import_utils
from import_utils import get_key
from large_delete_all import large_delete_all

today = timezone.localtime(timezone.now()).date()
earliest_year = (today - datetime.timedelta( days=106*365 )).year
latest_year = (today - datetime.timedelta( days=7*365 )).year

def date_from_value( s ):
	if isinstance(s, datetime.date):
		return s
	if isinstance(s, (float,int)):
		return datetime.date( *(xldate_as_tuple(s, import_utils.datemode)[:3]) )
	
	# Assume month, day, year format.
	mm, dd, yy = [int(v.strip()) for v in s.split( '/' )]
	
	# Correct for 2-digit year.
	for century in [0, 1900, 2000, 2100]:
		if earliest_year <= yy + century <= latest_year:
			yy += century
			break
	
	# Check if day and month are reversed.
	if mm > 12:
		dd, mm = mm, dd
		
	assert 1900 <= yy
		
	try:
		return datetime.date( year = yy, month = mm, day = dd )
	except Exception as e:
		print ( yy, mm, dd )
		raise e
		
def gender_from_str( s ):
	return 0 if s.lower().strip().startswith('m') else 1

def set_attributes( obj, attributes ):
	changed = False
	for key, value in attributes.items():
		if getattr(obj, key) != value:
			setattr(obj, key, value)
			changed = True
	return changed
	
def to_int_str( v ):
	try:
		return '{}'.format(v)
	except:
		pass
	return toUnicode(v)
		
def to_str( v ):
	return toUnicode(v).strip()

fnameDefault = r'EV\CyclingBC_28_Mar_2014_EV-Race-List.xls'
def init_ccn( fname = fnameDefault ):
	
	#large_delete_all( LicenseHolder )
	#large_delete_all( Team )
	
	tstart = datetime.datetime.now()
	
	fix_bad_license_codes()
	
	discipline_id = dict( (discipline, Discipline.objects.get(name=discipline)) for discipline in ['Road', 'Track', 'Cyclocross', 'MTB'] )

	effective_date = datetime.date.today()
	
	# Process the records in large transactions for efficiency.
	@transaction.atomic
	def process_ur_records( ur_records ):
		
		for i, ur in ur_records:
			try:
				date_of_birth	= date_from_value( ur.get('DOB', '') )
			except Exception as e:
				safe_print( 'Row {}: Invalid birthdate "{}" ({}) {}'.format( i, ur.get('DOB',''), ur, e ) )
				continue
				
			if not ur.get('License Numbers',''):
				safe_print( 'Row {}: Missing License Code '.format(i) )
				continue
			
			attributes = {
				'license_code':	to_int_str(ur.get('License Numbers','')),
				'last_name':	to_str(ur.get('Last Name','')),
				'first_name':	to_str(ur.get('First Name','')),
				'gender':		gender_from_str(to_str(ur.get('Sex','m'))),
				'date_of_birth':date_of_birth,
				'city':			to_str(ur.get('Member City','')),
				'state_prov':	to_str(ur.get('Member Province','')),
				'nationality':  to_str(ur.get('Nationality','')),
				'zip_postal':	to_str(get_key(ur,('ZipPostal','Zip''Postal','ZipCode','PostalCode','Zip Code','Postal Code',), None))
			}
			attributes = { a:v for a, v in attributes.items() if v is not None }
			
			if ur.get('Tag','').strip():
				attributes['existing_tag'] = to_int_str(ur.get('Tag','')).strip()
			
			try:
				lh = LicenseHolder.objects.get( license_code=attributes['license_code'] )
				if set_attributes(lh, attributes):
					lh.save()
			
			except LicenseHolder.DoesNotExist:
				lh = LicenseHolder( **attributes )
				lh.save()
			
			safe_print( '{i:>6}: {license:>8} {dob:>10} {uci_id} {last}, {first}, {city}, {state_prov}'.format(
				i=i,
				license=removeDiacritic(lh.license_code),
				dob=lh.date_of_birth.strftime('%Y/%m/%d'),
				uci_id=lh.uci_id,
				last=removeDiacritic(lh.last_name), first=removeDiacritic(lh.first_name),
				city=removeDiacritic(lh.city), state_prov=removeDiacritic(lh.state_prov) )
			)
			TeamHint.objects.filter( license_holder=lh ).delete()
			
			teams = [t.strip() for t in ur.get('Afiliates','').split(',')]
			if teams:
				last_team = None
				for t in teams:
					team = Team.objects.get_or_create( name=t )[0]
					last_team = team
				pcd = ur.get('Primary Cycling Discipline','').strip().lower()
				for id, d in discipline_id.items():
					if d.name.lower() in pcd:
						TeamHint( discipline=d, license_holder=lh, effective_date=effective_date, team=last_team )
						break
				
	suffix = 'CCN'
	ur_records = []
	wb = load_workbook( filename=fname, read_only=True, data_only=True  )
	
	for sheet_name in wb.sheetnames:
		if sheet_name.endswith(suffix):
			safe_print( 'Reading sheet: {}'.format(sheet_name) )
			ws = wb[sheet_name]
			break
	
	if not ws:
		safe_print( 'Cannot find sheet ending with "{}"'.format(suffix) )
		return
		
	for r, row in enumerate(ws.iter_rows()):
		if r == 0:
			# Get the header fields from the first row.
			fields = { col:toUnicode(f.value).strip() for col, f in enumerate(row) }
			safe_print( '\n'.join(fields.values()) )
			continue
		
		ur = { fields.get(col,''):cell.value for col, cell in enumerate(row) }
		ur_records.append( (r+1, ur) )
		if len(ur_records) == 3000:
			process_ur_records( ur_records )
			ur_records = []
			
	process_ur_records( ur_records )
	
	safe_print( 'Initialization in: ', datetime.datetime.now() - tstart )
