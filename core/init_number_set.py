import sys
import datetime
from openpyxl import load_workbook
from io import BytesIO
from django.db import transaction, IntegrityError
from django.db.models import Q
from . import import_utils
from .FieldMap import standard_field_map
from .import_utils import *
from .models import *

def init_number_set( numberSetId, worksheet_name='', worksheet_contents=None, message_stream=sys.stdout ):
	
	tstart = datetime.datetime.now()

	if message_stream == sys.stdout or message_stream == sys.stderr:
		def ms_write( s ):
			message_stream.write( removeDiacritic(s) )
	else:
		def ms_write( s ):
			message_stream.write( '{}'.format(s) )
	
	try:
		number_set = NumberSet.objects.get( pk=numberSetId )
	except NumberSet.DoesNotExist:
		ms_write( '**** Cannot find NumberSet\n' )
		return
	
	ifm = standard_field_map()
	
	# Process the records in large transactions for efficiency.
	def process_ur_records( ur_records ):
		for i, ur in ur_records:
			v = ifm.finder( ur )
			
			bib				= v('bib','')
			try:
				bib = int(bib)
			except ValueError:
				ms_write( '**** Row {:>6}: invalid Bib: {}"\n'.format(i, bib) )
				continue
			
			success = True
			license_holder = None
			for f, k, v in ( ('UCIID', 'uci_id', to_uci_id(v('uci_id', None))), ('License', 'license_code', to_int_str(v('license_code', '')).upper().strip()) ):
				if not v:
					continue
					
				license_holders = list( LicenseHolder.objects.filter( **{k:v} ) )
				if not license_holders:
					ms_write( '**** Row {:>6}: cannot find LicenseHolder from {}: "{}"\n'.format(i, f, v) )
				elif len(license_holders) != 1:
					ms_write( '**** Row {:>6}: Ignoring.  Multiple LicenseHolders found {}: "{}"\n'.format(i, f, v) )
					for c, license_holder in enumerate(license_holders, 1):
						ms_write(
							'     {c:>3}. {license_code:>16} {uci_id} {last_name}, {first_name}, {city}, {state_prov}\n'.format(
								c=c,
								license_code=license_holder.license_code,
								uci_id=license_holder.uci_id,
								last_name=license_holder.last_name,
								first_name=license_holder.first_name,
								city=license_holder.city, state_prov=license_holder.state_prov,
							)
						)
				else:
					license_holder = license_holders[0]
					success = True
				break

			if not success:
				continue

			if license_holder:
				if not number_set.assign_bib( license_holder, bib ):
					ms_write( '**** Row {:>6}: bib={} cannot be assigned.  The bib must be allowed by the NumberSet ranges.\n'.format(
						i, bib) )
				else:	
					ms_write(
						'Row {i:>6}: {bib:>4} {license_code:>16} {uci_id} {last_name}, {first_name}, {city}, {state_prov}\n'.format(
							i=i,
							bib=bib,
							license_code=license_holder.license_code,
							uci_id=license_holder.uci_id,
							last_name=license_holder.last_name,
							first_name=license_holder.first_name,
							city=license_holder.city, state_prov=license_holder.state_prov,
						)
					)
			else:
				number_set.set_lost( bib )
				ms_write(
					'Row {i:>6}: {bib:>4} Lost\n'.format(
						i=i,
						bib=bib,
					)
				)
			
	
	sheet_name = None
	if worksheet_contents is not None:
		wb = load_workbook( filename = BytesIO(worksheet_contents), read_only=True, data_only=True )
	else:
		try:
			fname, sheet_name = worksheet_name.split('$')
		except Exception:
			fname = worksheet_name
		wb = load_workbook( filename = fname, read_only=True, data_only=True )
	
	try:
		sheet_name = sheet_name or wb.sheetnames[0]
		ws = wb[sheet_name]
		ms_write( 'Reading sheet "{}"\n'.format(sheet_name) )
	except Exception as e:
		ms_write( 'Cannot find sheet "{} ({})"\n'.format(sheet_name, e) )
		return
		
	ur_records = []
	for r, row in enumerate(ws.iter_rows()):
		if r == 0:
			# Get the header fields from the first row.
			fields = ['{}'.format(f.value).strip() for f in row]
			ifm.set_headers( fields )
			
			expected_fields = {'bib', 'license_code', 'uci_id'}
			
			ms_write( 'Header Row:\n' )
			for col, f in enumerate(fields, 1):
				name = ifm.get_name_from_alias( f )
				if name in expected_fields:
					ms_write( '        {}. {} --> {}\n'.format(col, f, name) )
				else:
					ms_write( '        {}. ****{} (Ignored)\n'.format(col, f) )
					if name in ifm:
						del ifm[name]
			
			if 'bib' not in ifm:
				ms_write( 'Header Row must contain "Bib"\n' )
				return
						
			if sum( int(f in ifm) for f in ('uci_id', 'license_code') ) == 0:
				ms_write( 'Header Row must contain either "UCIID" or "License"\n' )
				return
				
			ms_write( '\n' )
			continue	
			
			fields = {col:'{}'.format(f.value).strip() for col, f in enumerate(row)}
			ms_write( 'Header Row:\n' )
			for f in fields.values():
				ms_write( '   {}\n'.format(f) )
			
			fields_lower = set( str(f).lower() for f in fields.values() )
			if not any( n.lower() in fields_lower for n in license_col_names ):
				ms_write( 'License column not found in Header Row.  Aborting.\n' )
				return
			if not any( n.lower() in fields_lower for n in bib_col_names ):
				ms_write( 'Bib column not found in Header Row.  Aborting.\n' )
				return
			continue
			
		ur_records.append( (r+1, [v.value for v in row]) )
		if len(ur_records) == 1000:
			process_ur_records( ur_records )
			ur_records = []
			
	process_ur_records( ur_records )
	
	ms_write( '\n' )
	ms_write( 'Initialization in: {}\n'.format(datetime.datetime.now() - tstart) )
