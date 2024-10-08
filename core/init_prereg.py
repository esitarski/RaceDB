import io
import re
import sys
import datetime
import operator
from fnmatch import fnmatch
from openpyxl import load_workbook
from django.db import transaction, IntegrityError
from django.db.models import Q

from .large_delete_all import large_delete_all
from .FieldMap import standard_field_map, normalize
from .import_utils import *
from .models import *

class TimeTracker:
	def __init__( self ):
		self.startTime = None
		self.curLabel = None
		self.totals = defaultdict( float )
		
	def start( self, label ):
		t = datetime.datetime.now()
		if self.curLabel and self.startTime:
			self.totals[self.curLabel] += (t - self.startTime).total_seconds()
		self.curLabel = label
		self.startTime = t
		
	def end( self ):
		self.start( None )
		
	def __repr__( self ):
		s = []
		for lab, t in sorted( self.totals.items(), key=operator.itemgetter(1), reverse=True ):
			s.append( '{:<50}: {:6.2f}'.format(lab, t) )
		s.append( '' )
		return '\n'.join( s )

def init_prereg(
		competition_name='', worksheet_name='', assign_missing_bibs=False, clear_existing=False,
		competitionId=None, worksheet_contents=None, message_stream=sys.stdout ):

	tt = TimeTracker()
		
	team_lookup = TeamLookup()
		
	tstart = datetime.datetime.now()
	today = datetime.date.today()

	if message_stream == sys.stdout or message_stream == sys.stderr:
		def ms_write( s, flush=False ):
			message_stream.write( removeDiacritic(s) )
	else:
		def ms_write( s, flush=False ):
			message_stream.write( '{}'.format(s) )
			sys.stdout.write( removeDiacritic(s) )
			if flush:
				sys.stdout.flush()
	
	fix_bad_license_codes()
	
	if competitionId is not None:
		competition = Competition.objects.get( pk=competitionId )
	else:
		try:
			competition = Competition.objects.get( name=competition_name )
		except Competition.DoesNotExist:
			ms_write( '**** Cannot find Competition: "{}"\n'.format(competition_name) )
			return
		except Competition.MultipleObjectsReturned:
			ms_write( '**** Found multiple Competitions matching: "{}"\n'.format(competition_name) )
			return
	
	optional_events = { normalize(event.name):event for event in competition.get_events() if event.optional }
	pattern_optional_events = defaultdict( list )
	
	# Check for duplicate event names.
	event_name_count = defaultdict( int )
	for event in competition.get_events():
		if event.optional:
			event_name_count[event.name] += 1
	for event_name, count in event_name_count.items():
		if count > 1:
			ms_write( '**** Error: Duplicate Optional Event Name: "{}".  Perferences to Optional Events may not work properly.\n'.format(event_name) )
	
	role_code = {}
	for role_type, roles in Participant.COMPETITION_ROLE_CHOICES:
		role_code.update( { '{}'.format(name).lower().replace(' ','').replace('.',''):code for code, name in roles } )
		
	# Construct a cache to find categories quicker.
	category_code_gender_suffix = re.compile( r'\(Open\)$|\(Men\)$|\(Women\)$' )
	cat_gender = defaultdict( list )
	category_numbers_set = {None: set()}
	for category in Category.objects.filter( format=competition.category_format ).order_by('gender', 'code'):
		cat_gender[category.code].append( (category, category.gender) )
		cn = competition.get_category_numbers( category )
		category_numbers_set[category] = cn.get_numbers() if cn else set()
	
	def get_category( category_code_search, gender_search ):
		category_code_search = category_code_gender_suffix.sub( '', category_code_search ).strip()
		for category, gender in cat_gender.get(category_code_search, []):
			if gender_search == gender or gender == 2:
				return category
		return None
		
	times = defaultdict(float)
	
	has_legal_entity = competition.legal_entity
	ifm = standard_field_map( exclude = ([] if has_legal_entity else ['waiver']) )
	if has_legal_entity:
		waiver_signed_date = competition.legal_entity.waiver_expiry_date + datetime.timedelta(days=1)
	
	# Process the records in large transactions for efficiency.
	def process_ur_records( ur_records ):
		for i, ur in ur_records:
			
			tt.start( 'get_fields_from_row' )
		
			v = ifm.finder( ur )
			date_of_birth	= v('date_of_birth', None)
			try:
				date_of_birth = date_from_value(date_of_birth)
			except Exception as e:
				ms_write( 'Row {}: Ignoring birthdate (must be YYYY-MM-DD) "{}" ({}) {}'.format(
					i, date_of_birth, ur, e)
				)
				date_of_birth = None
			date_of_birth 	= date_of_birth if date_of_birth != invalid_date_of_birth else None
			
			uci_code = to_str(v('uci_code', None))
			# If no date of birth, get it from the UCI code.
			if not date_of_birth and uci_code and not uci_code.isdigit():
				try:
					date_of_birth = datetime.date( int(uci_code[3:7]), int(uci_code[7:9]), int(uci_code[9:11]) )
				except Exception:
					pass
			
			# If no date of birth, make one up based on the age.
			age = to_int(v('age', None))
			if not date_of_birth and age:
				date_of_birth = date_of_birth_from_age( age, today )
				year_only_dob = True
			else:
				year_only_dob = False
			
			# As a last resort, pick the default DOB
			date_of_birth 	= date_of_birth or invalid_date_of_birth
			
			license_code	= (to_int_str(v('license_code', '')) or '').upper().strip() or None
			last_name		= to_str(v('last_name',''))
			first_name		= to_str(v('first_name',''))
			name			= to_str(v('name',''))
			if not name:
				name = ' '.join( n for n in [first_name, last_name] if n )
			
			gender			= to_str(v('gender',''))
			gender			= gender_from_str(gender) if gender else None
			
			email			= to_str(v('email', None))
			phone			= to_phone(v('phone', None))
			city			= to_str(v('city', None))
			state_prov		= to_str(v('state_prov', None))
			zip_postal		= to_str(v('zip_postal', None))
			if zip_postal:
				zip_postal = validate_postal_code( zip_postal )
			
			preregistered	= to_bool(v('preregistered', True))
			paid			= to_bool(v('paid', None))
			bib				= (to_int(v('bib', None)) or None)
			bib_auto		= (bib is None and '{}'.format(v('bib','')).lower() == 'auto')
			tag				= to_int_str(v('tag', None))
			note		 	= to_str(v('note', None))
			team_name		= to_str(v('team', None))
			club_name		= to_str(v('club', None))
			if not team_name:
				team_name = club_name
			category_code   = to_str(v('category_code', None))
			
			est_kmh			= to_float(v('est_kmh', None))
			est_mph			= to_float(v('est_mph', None))
			if est_mph is not None:
				est_kmh = 1.609344 * est_mph
				est_mph = None
			seed_option		= to_str(v('seed_option', None))
			if seed_option is not None:
				seed_option = seed_option.lower()
				if 'early' in seed_option:
					seed_option = 0
				elif 'late' in seed_option:
					seed_option = 2
				elif 'last' in seed_option:
					seed_option = 3
				else:
					seed_option = 1
			
			uci_id			= to_uci_id(v('uci_id', None))
			nation_code		= to_str(v('nation_code', None))
			
			emergency_contact_name = to_str(v('emergency_contact_name', None))
			emergency_contact_phone = to_phone(v('emergency_contact_phone', None))
			emergency_medical = to_str(v('emergency_medical', None))
			
			participant_optional_events = []
			for pattern, events in pattern_optional_events.items():
				included = to_bool(v(pattern, False))
				participant_optional_events.extend( (event, included) for event in events )
			
			race_entered    = to_str(v('race_entered', None))
			
			role			= role_code.get( v('role','').lower().replace(' ', '').replace('.', ''), None )
			waiver			= to_bool(v('waiver',None))
			license_checked	= to_bool(v('license_checked',None))
			
			#------------------------------------------------------------------------------
			# Get LicenseHolder.
			#
			license_holder = None
			#with transaction.atomic():
			if True:
				tt.start( 'get_license_holder' )
			
				if uci_id:
					license_holder = LicenseHolder.objects.filter( uci_id=uci_id ).first()
					
				if not license_holder and license_code and license_code.upper() != 'TEMP':
					try:
						license_holder = LicenseHolder.objects.get( license_code=license_code )
					except LicenseHolder.DoesNotExist:
						ms_write( '**** Row {}: cannot find LicenceHolder from LicenseCode: {}, Name="{}"\n'.format(
							i, license_code, name) )
						continue
				elif not license_holder:
					# No license code.  Try to find the participant by last/first name, [date_of_birth] and [gender].
					# Case insensitive comparison, accents ignored for names.
					q = Q( search_text__startswith=utils.get_search_text([last_name, first_name]) )
					if date_of_birth and date_of_birth != invalid_date_of_birth:
						if year_only_dob:
							q &= Q( date_of_birth__year=date_of_birth.year )
						else:
							q &= Q( date_of_birth=date_of_birth )
					if gender is not None:
						q &= Q( gender=gender )
					
					try:
						license_holder = LicenseHolder.objects.get( q )
					except LicenseHolder.DoesNotExist:
						# No name match.
						# Create a temporary license holder.
						try:
							license_holder = LicenseHolder(
								**{ attr:value for attr, value in {
										'license_code':'TEMP',
										'last_name':last_name,
										'first_name':first_name,
										'gender':gender,
										'date_of_birth':date_of_birth,
										'uci_code':uci_code,
										'nation_code':nation_code,
										'uci_id':uci_id,
										'email':email,
										'phone':phone,
										'city':city,
										'state_prov':state_prov,
										'zip_postal':zip_postal,
										'emergency_contact_name':emergency_contact_name,
										'emergency_contact_phone':emergency_contact_phone,
										'emergency_medical':emergency_medical,
										'existing_tag':tag if competition.use_existing_tags else None,
									}.items() if value is not None
								}
							)
							truncate_char_fields(license_holder).save()
						except Exception as e:
							ms_write( '**** Row {}: New License Holder Exception: {}, Name="{}"\n'.format(
									i, e, name,
								)
							)
							continue
					except LicenseHolder.MultipleObjectsReturned:
						ms_write( '**** Row {}: found multiple LicenceHolders matching "Last, First" Name="{}"\n'.format(
								i, name,
							)
						)
						continue
				
				# Update the license_holder record with all new information.
				tt.start( 'update_license_holder' )
				if set_attributes( license_holder, {
						'date_of_birth':date_of_birth
							if not year_only_dob and date_of_birth != invalid_date_of_birth
							else None,
						'uci_code':uci_code,
						'email':email,
						'phone':phone,
						'city':city,
						'state_prov':state_prov,
						'zip_postal':zip_postal,
						'nation_code':nation_code,
						'uci_id':uci_id,
						'emergency_contact_name':emergency_contact_name,
						'emergency_contact_phone':emergency_contact_phone,
						'emergency_medical':emergency_medical,
						'existing_tag':tag if competition.use_existing_tags else None,
					} ):
					try:
						truncate_char_fields(license_holder).save()
						truncate_char_fields( license_holder )
					except Exception as e:
						ms_write( '**** Row {}: Update License Holder Exception: {}, Name="{}"\n'.format(
								i, e, name,
							)
						)
						continue
				
				#------------------------------------------------------------------------------
				# Get Category.  Open categories will match either Gender.
				#
				category = None
				if category_code:
					tt.start( 'get_category_from_code' )
					category = get_category( category_code, license_holder.gender )
					if category is None:
						ms_write( '**** Row {}: cannot match Category (ignoring): "{}" Name="{}"\n'.format(
							i, category_code, name,
						) )
				
				#------------------------------------------------------------------------------
				# Get Team
				#
				tt.start( 'get_team' )
				team = None
				if Team.is_independent_name(team_name):
					team = None
				else:
					if team_name not in team_lookup:
						msg = 'Row {:>6}: Added team: {}\n'.format(
							i, team_name,
						)
						ms_write( msg )
					team = team_lookup[team_name]
				
				#------------------------------------------------------------------------------
				tt.start( 'get_participant' )
				participant_keys = { 'competition': competition, 'license_holder': license_holder, }
				if category is not None:
					participant_keys['category'] = category
				
				participant_created = False
				try:
					participant = Participant.objects.get( **participant_keys )
				except Participant.DoesNotExist:
					participant = Participant( **participant_keys )
					participant_created = True
				except Participant.MultipleObjectsReturned:
					ms_write( '**** Row {}: found multiple Participants for this license_holder, Name="{}".\n'.format(
						i, name,
					) )
					continue
				
				# First get the defaults based on previous hints and categories.
				tt.start( 'init_participant_defaults' )
				participant.init_default_values()
				
				# Now, override default values with specified ones.
				tt.start( 'override_participant_defaults' )
				for attr, value in (
						('category',category),
						('bib',bib), ('tag',tag), ('note',note),
						('preregistered',preregistered), ('paid',paid), 
						('seed_option',seed_option), ('est_kmh',est_kmh), 
						('role',role),
						('license_checked',license_checked),
					):
					if value is not None:
						setattr( participant, attr, value )
				if team_name:
					participant.team = team

				# Ensure the existing bib number is compatible with the category.
				tt.start( 'update_bib_new_category' )
				bib = participant.update_bib_new_category( category_numbers_set[participant.category] )
				
				# Auto-assign the bib if there isn't one already.
				if not bib and (assign_missing_bibs or bib_auto):
					tt.start( 'get_bib_auto' )
					
					# Retrieve the existing bib if it exists for this license_holder and category.
					if competition.number_set:
						bib = competition.number_set.get_bib( competition, license_holder, category, category_numbers_set[participant.category] )
						has_existing_bib = bool( bib )
					
					# If no existing bib, allocate the next available one.
					if not bib:
						bib = participant.get_bib_auto()
						if bib and competition.number_set:
							allocated_bib = competition.number_set.assign_bib( participant.license_holder, bib )
							
					if not bib:
						ms_write( '**** Row {}: Error={}\nCategory={} Name="{}"\n'.format(
							i, 'Auto-allocate bib failed: Category Numbers undefined or full.',
							category_code, name,
						) )
				
				# If we have an assigned bib, ensure it is respected.
				tt.start( 'validate_bib' )
				if bib and bib in category_numbers_set[participant.category]:
					participant.bib = bib
					if competition.number_set:
						competition.number_set.assign_bib( participant.license_holder, participant.bib )

				# Final check for participant bib integrity.
				if participant.bib and participant.bib not in category_numbers_set[participant.category]:
					participant.bib = None
				
				participant.preregistered = True
				
				tt.start( 'participant_save' )
				try:
					truncate_char_fields(participant).save()
				except IntegrityError as e:
					ms_write( '**** Row {}: Error={}\nBib={} Category={} License={} Name="{}"\n'.format(
						i, e,
						bib, category_code, license_code, name,
					) )
					success, integrity_error_message, conflict_participant = participant.explain_integrity_error()
					if success:
						ms_write( '{}\n'.format(integrity_error_message) )
						ms_write( '{}\n'.format(conflict_participant) )
					continue
				
				tt.start( 'add_to_default_optional_events' )
				participant.add_to_default_optional_events()
				
				if participant_optional_events:
					participant_optional_events = {
						event:(included and event.could_participate(participant))
						for event, included in participant_optional_events
					}
					option_included = { event.option_id:included for event, included in participant_optional_events.items() }
					ParticipantOption.sync_option_ids( participant, option_included )
					override_events_str = ' ' + ', '.join(
						'"{}"={}'.format(event.name, included)
							for event, included in sorted(
								participant_optional_events.items(), key=lambda event_included: event_included[0].date_time
							)
					)
				else:
					override_events_str = ''
				
				if waiver is not None:
					tt.start( 'waiver' )
					if waiver:
						participant.sign_waiver_now( waiver_signed_date )
					else:
						participant.unsign_waiver_now()
						
				if not category and not participant_optional_events:
					tt.start( 'add_other_categories' )
					participant.add_other_categories()
				
				tt.end()
				
				ms_write( '{created:} Row {row:>6}: {uci:>11} {license:>12} bib={bib:>4}, {lname}, {fname}, {city}, {state_prov} {ov}\n'.format(
						created='*' if participant_created else ' ',
						row=i,
						uci=license_holder.uci_id or '',
						license=(license_holder.license_code or '').replace(' ', ''),
						bib=participant.bib or '',
						lname=license_holder.last_name,
						fname=license_holder.first_name,
						city=license_holder.city,
						state_prov=license_holder.state_prov,
						ov=override_events_str,
					)
				)
	
	sheet_name = None
	if worksheet_contents is not None:
		wb = load_workbook( filename = io.BytesIO(worksheet_contents), read_only=True, data_only=True )
	else:
		try:
			fname, sheet_name = worksheet_name.split('$')
		except Exception:
			fname = worksheet_name
		wb = load_workbook( filename = fname, read_only=True, data_only=True )
	
	try:
		sheet_name = sheet_name or wb.sheetnames[0]
		ws = wb[sheet_name]
		ms_write( 'Reading sheet "{}"\n'.format(sheet_name) )
	except Exception:
		ms_write( 'Cannot find sheet "{}"\n'.format(sheet_name) )
		return
		
	ur_records = []
	
	for r, row in enumerate(ws.iter_rows()):
		if r == 0:
			# Get the header fields from the first row.
			fields = ['{}'.format(v.value).strip() for v in row]
			
			# Add all Optional Event patterns.
			for field in fields:
				try:
					pattern = normalize( field )
				except Exception:
					continue
				for event_name, event in optional_events.items():
					if fnmatch(event_name, pattern):
						pattern_optional_events[pattern].append( event )
			for pattern in pattern_optional_events.keys():
				ifm.set_aliases( pattern, (pattern,) )
			
			ifm.set_headers( fields )
			ms_write( 'Header Row:\n' )
			for col, f in enumerate(fields, 1):
				if f.lower().strip() in optional_events:
					ms_write( '        {}. {} (Optional Event)\n'.format(col,f) )
				else:
					name = ifm.get_name_from_alias( f )
					if name is not None:
						ms_write( '        {}. {} --> {}\n'.format(col, f, name) )
					else:
						ms_write( '        {}. ****{} (Ignored)\n'.format(col, f) )
			
			fields_lower = [f.lower() for f in fields]
			if clear_existing or 'bib' in ifm:
				ms_write( 'Recording previous license checks...\n' )
				if competition.report_label_license_check:
					tt.start( 'license_check_state_refresh' )
					LicenseCheckState.refresh()
	
				ms_write( 'Clearing existing Participants...\n' )
				tt.start( 'clearing_existing_participants' )
				large_delete_all( Participant, Q(competition=competition) )
			
			if 'license_code' not in ifm and 'uci_id' not in ifm:
				ms_write( 'Header Row must contain one of (or both) License or UCIID.  Aborting.\n' )
				return
			
			ms_write( '\n' )
			continue
			
		ur_records.append( (r+1, [v.value for v in row]) )
		if len(ur_records) == 300:
			process_ur_records( ur_records )
			ur_records[:] = []
			
	process_ur_records( ur_records )
	
	ms_write( '\n' )
	for section, total in sorted( times.items(), key = operator.itemgetter(1), reverse=True ):
		ms_write( '{}={:.6f}\n'.format(section, total) )
	ms_write( '\n' )
	ms_write( 'Initialization in: {}\n'.format(datetime.datetime.now() - tstart) )
	ms_write( '\n' )
	ms_write( tt.__repr__() )
