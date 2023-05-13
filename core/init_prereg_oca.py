import sys
import datetime
from openpyxl import load_workbook
from html.parser import HTMLParser
from html import unescape
from collections import namedtuple
from django.db import transaction, IntegrityError
from django.db.models import Q

from .models import *
from .import_utils import *
from .large_delete_all import large_delete_all
from .FieldMap import standard_field_map

datemode = None

today = datetime.date.today()
earliest_year = (today - datetime.timedelta( days=106*365 )).year
latest_year = (today - datetime.timedelta( days=7*365 )).year

def date_from_value( s ):
	if isinstance(s, datetime.date):
		return s
	if isinstance(s, (float, int)):
		return datetime.date( *(xldate_as_tuple(s, datemode)[:3]) )
	
	# Assume month, day, year format.
	mm, dd, yy = [int(v.strip()) for v in s.split( '/' )]
	
	# Correct for 2-digit year.
	for century in [0, 1900, 2000, 2100]:
		if earliest_year <= yy + century <= latest_year:
			yy += century
			break
	
	# Check if day and month are reversed.
	if mm > 12:
		dd, mm = mm, dd
		
	assert 1900 <= yy
		
	try:
		return datetime.date( year = yy, month = mm, day = dd )
	except Exception as e:
		safe_print( yy, mm, dd )
		raise e
		
def gender_from_str( s ):
	return 0 if s.lower().strip().startswith('m') else 1

def set_attributes( obj, attributes ):
	changed = False
	for key, value in attributes.items():
		if getattr(obj, key) != value:
			setattr(obj, key, value)
			changed = True
	return changed
	
def to_int_str( v ):
	try:
		return '{}'.format(int(v))
	except Exception:
		pass
	return '{}'.format(v)
		
def to_str( v ):
	if v is None:
		return v
	return '{}'.format(v)
	
def to_bool( v ):
	if v is None:
		return None
	s = '{}'.format(v)
	return s[:1] in 'YyTt1'

def to_int( v ):
	if v is None:
		return None
	try:
		return int(v)
	except Exception:
		return None
		
def to_tag( v ):
	if v is None:
		return None
	return '{}'.format(v).split('.')[0]

def init_prereg_oca( competition_name, worksheet_name, clear_existing ):
	global datemode
	
	tstart = datetime.datetime.now()
	
	html_parser = HTMLParser()

	try:
		competition = Competition.objects.get( name=competition_name )
	except Competition.DoesNotExist:
		safe_print( 'Cannot find Competition: "{}"'.format(competition_name) )
		return
	except Competition.MultipleObjectsReturned:
		safe_print( 'Found multiple Competitions matching: "{}"'.format(competition_name) )
		return
	
	if clear_existing:
		Participant.objects.filter(competition=competition).delete()
	
	ifm = standard_field_map()
	competition_categories = list( competition.category_format.category_set.all() )
	for c in competition_categories:
		ifm.set_aliases( str(c.id), [c.code] )
	
	# Process the records in large transactions for efficiency.
	@transaction.atomic
	def process_ur_records( ur_records ):
		
		for i, fields in ur_records:
			v = ifm.finder( fields )
			
			bib				= to_int(v.get_value('bibnum', fields, None))
			license_code	= 'ON' + (to_int_str(v.get_value('license', fields, '')) or to_int_str(v.get_value('licence', fields, '<missing license code>')))
			uci_id			= to_str(v.get_value('uci_id', fields, '<missing uci id>'))
			name			= to_str(v.get_value('name', fields,'')).strip()
			team_name		= to_str(v.get_value('Team', fields, None))
			preregistered	= to_bool(v.get_value('preregistered', fields, True))
			paid			= to_bool(v.get_value('paid', fields, None))
			category_code   = to_str(v.get_value('category', fields, None))
			
			first_name		= v.get_value('first_name', fields, None)
			last_name		= v.get_value('last_name', fields, None)
			date_of_birth	= v.get_value('date_of_birth', fields, None)
			
			tag				= v.get_value('tag', fields, None)

			emergency_contact_name = to_str(v.get_value('emergency_contact_name', fields, None))
			emergency_contact_phone = to_str(v.get_value('emergency_contact_phone', fields, None))
			
			license_holder = LicenseHolder.objects.filter( uci_id=uci_id ).first()
			
			if not license_holder:
				try:
					license_holder = LicenseHolder.objects.get( license_code=license_code )
				except LicenseHolder.DoesNotExist:
					pass
			
			if not license_holder and first_name and last_name and date_of_birth:
				try:
					license_holder = LicenseHolder.objects.get( first_name=first_name, last_name=last_name, date_of_birth=date_of_birth )
				except Exception as e:
					pass				
			
			if not license_holder:
				safe_print( 'Row {}: cannot find LicenceHolder from UCI ID ({}), LicenceCode ({}) or First/Last/DOB'.format(i, uci_id, license_code) )
				continue
					
			do_save = False
			for fname, fvalue in (  ('emergency_contact_name',emergency_contact_name),
									('emergency_contact_phone',emergency_contact_name) ):
				if fvalue and getattr(license_holder, fname) != fvalue:
					setattr( license_holder, fname, fvalue )
					do_save = True
			
			if do_save:
				license_holder.save()
				
			team = None
			if team_name is not None:
				try:
					team = Team.objects.get(name=team_name)
				except Team.DoesNotExist:
					safe_print( 'Row {}: unrecognized Team name (ignoring): "{}"'.format(
						i, team_name,
					) )
				except Team.MultipleObjectsReturned:
					safe_print( 'Row {}: multiple Teams match name (ignoring): "{}"'.format(
						i, team_name,
					) )
			participant.team = team

			# Check which speecific categories are entered.
			category = None
			if category_code is not None:
				try:
					category = Category.objects.get( format=competition.category_format, code=category_code )
				except Category.DoesNotExist:
					safe_print( 'Row {}: unrecognized Category code (ignoring): "{}"'.format(
						i, category_code,
					) )
				except Category.MultipleObjectsReturned:
					safe_print( 'Row {}: multiple Categories match code (ignoring): "{}"'.format(
						i, category_code,
					) )

			categories_entered = [category] if category else []
			if not categories_entered:
				for c in competition_categories:
					in_category = v.get_value( str(c.id), fields, None )
					if in_category:
						categories_entered.append( c )

`			if not categories_entered:
				categories_entered = [None]

			for category in categories_entered:
				query_values = dict( competition=competition, license_holder=license_holder )
				if category:
					query_values['category'] = category
				try:
					participant = Participant.objects.get( **query_values )
				except Participant.DoesNotExist:
					participant = Participant( **query_values )
				except Participant.MultipleObjectsReturned:
					safe_print( 'Row {}: found multiple Participants for this competition and license_holder.'.format(
							i,
					) )
					continue
				
				for attr, value in [ ('preregistered',preregistered), ('paid',paid), ('bib',bib), ]:
					if value is not None:
						setattr( participant, attr, value )
	v			
				try:
					participant.save()
				except IntegrityError as e:
					safe_print( 'Row {}: Error={}\nBib={} Category={} License={} Name={}'.format(
						i, e,
						bib, category_code, license_code, name,
					) )
					continue
			
				if not category:
					participant.add_to_default_optional_events()
			
			safe_print( '{:>6}: {:>8} {:>10} {}, {}, {}, {}, {}'.format(
					i,
					license_holder.license_code, license_holder.date_of_birth.strftime('%Y/%m/%d'),
					license_holder.uci_id,
					license_holder.last_name, license_holder.first_name,
					license_holder.city, license_holder.state_prov
				)
			)
	
	try:
		fname, sheet_name = worksheet_name.split('$')
	except Exception:
		fname = worksheet_name
		sheet_name = None
		
	wb = load_workbook( filename=fname, read_only=True, data_only=True )
	
	try:
		sheet_name = sheet_name or wb.sheetnames[0]
		ws = wb[sheet_name]
		ms_write( 'Reading sheet "{}"\n'.format(sheet_name) )
	except Exception:
		ms_write( 'Cannot find sheet "{}"\n'.format(sheet_name) )
		return
		
	ur_records = []
	for r, row in enumerate(ws.iter_rows()):
		if r == 0:
			# Get the header fields from the first row.
			fields = [unescape('{}'.format(v.value).strip()).replace('-','_').replace('#','').strip().replace('4', 'four').replace(' ','_').lower() for v in row]
			
			ifm.set_headers( fields )
			
			safe_print( '\n'.join( ifm.get_names() ) )
			continue
			
		record = tuple( cell.value for cell in row )
		ur_records.append( (r+1, record) )
		if len(ur_records) == 1000:
			process_ur_records( ur_records )
			ur_records = []
			
	process_ur_records( ur_records )
	
	safe_print( 'Initialization in: ', datetime.datetime.now() - tstart )
