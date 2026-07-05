import io
import sys
import datetime
from openpyxl import load_workbook
from . import import_utils
from .FieldMap import standard_field_map
from .import_utils import *
from .models import *

def init_ranking( rankingId, worksheet_name='', worksheet_contents=None, message_stream=sys.stdout ):
	
	tstart = datetime.datetime.now()

	if message_stream == sys.stdout or message_stream == sys.stderr:
		def ms_write( s ):
			message_stream.write( removeDiacritic(s) )
	else:
		def ms_write( s ):
			message_stream.write( '{}'.format(s) )
	
	try:
		ranking = Ranking.objects.get( pk=rankingId )
	except Ranking.DoesNotExist:
		ms_write( '**** Cannot find Ranking\n' )
		return
	
	sheet_name = None
	if worksheet_contents is not None:
		wb = load_workbook( filename = io.BytesIO(worksheet_contents), read_only=True, data_only=True )
	else:
		try:
			fname, sheet_name = worksheet_name.split('$')
		except Exception:
			fname = worksheet_name
		wb = load_workbook( filename = fname, read_only=True, data_only=True )
	
	try:
		sheet_name = sheet_name or wb.sheetnames[0]
		ws = wb[sheet_name]
		ms_write( 'Reading sheet "{}"\n'.format(sheet_name) )
	except Exception:
		ms_write( 'Cannot find sheet "{}"\n'.format(sheet_name) )
		return
	
	ifm = standard_field_map()
	
	to_add = []
	for i, row in enumerate(ws.iter_rows()):
		if i == 0:
			# Get the header fields from the first row.
			fields = ['{}'.format(f.value).strip() for f in row]
			ifm.set_headers( fields )
			
			expected_fields = {'license_code', 'uci_id', 'rank', 'points'}
			
			ms_write( 'Header Row:\n' )
			for col, f in enumerate(fields, 1):
				name = ifm.get_name_from_alias( f )
				if name in expected_fields:
					ms_write( '        {}. {} --> {}\n'.format(col, f, name) )
				else:
					ms_write( '        {}. ****{} (Ignored)\n'.format(col, f) )
					if name in ifm:
						del ifm[name]
			
			if sum( int(f in ifm) for f in ('uci_id', 'license_code') ) == 0:
				ms_write( 'Header Row must contain "UCIID" or "License" or both\n' )
				return
				
			ms_write( '\n' )
			continue	
		
		values = [v.value for v in row]
		v = ifm.finder( values )
		
		uci_id = v('uci_id', None)
		if uci_id:
			if isinstance(uci_id, float):
				uci_id = str(int(uci_id))
			uci_id_error = get_uci_id_error( uci_id )
			if uci_id_error:
				ms_write( '**** Row {:>6}: Ignoring. UCI ID error: {} ({})\n'.format(i, uci_id_error, uci_id) )
				continue
				
		license_code = v('license_code', None)
		if uci_id is None and license_code is None:
			ms_write( '**** Row {:>6}: Ignoring. A UCI ID and/or a License Code must be present\n'.format(i) )
			continue
		rank = v('rank', None)
		if rank is None:
			ms_write( '**** Row {:>6}: Ignoring. missing rank\n'.format(i) )
		if isinstance(rank, float):
			rank = int( rank )
		if isinstance( rank, str ):
			try:
				rank = int( rank )
			except Exception:
				ms_write( '**** Row {:>6}: Ignoring. Invalid rank (must be integer)\n'.format(i) )
				continue
				
		if not rank or not str(rank).isdigit():
			ms_write( '**** Row {:>6}: Ignoring. Rank is invalid or missing\n'.format(i) )
			coninue
		
		points = v('points', None)
		to_add.append( RankingEntry(ranking=ranking, uci_id=uci_id, license_code=license_code, rank=rank, points=points) )

	ranking.rankingentry_set.all().delete()
	RankingEntry.objects.bulk_create( to_add )
	
	ms_write( '\n' )
	ms_write( 'Initialization in: {}\n'.format(datetime.datetime.now() - tstart) )
