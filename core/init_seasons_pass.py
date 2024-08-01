import sys
import datetime
from openpyxl import load_workbook
from io import BytesIO
from django.db import transaction, IntegrityError
from django.db.models import Q

from .large_delete_all import large_delete_all
from . import import_utils
from .FieldMap import standard_field_map
from .import_utils import *
from .models import *

def init_seasons_pass( seasonsPassId, worksheet_name='', worksheet_contents=None, message_stream=sys.stdout, clear_existing=False ):
	
	tstart = datetime.datetime.now()

	if message_stream == sys.stdout or message_stream == sys.stderr:
		def ms_write( s ):
			message_stream.write( removeDiacritic(s) )
	else:
		def ms_write( s ):
			message_stream.write( '{}'.format(s) )
			sys.stdout.write( removeDiacritic(s) )
	
	try:
		seasons_pass = SeasonsPass.objects.get( pk=seasonsPassId )
	except SeasonsPass.DoesNotExist:
		ms_write( '**** Cannot find SeasonsPass\n' )
		return
		
	if clear_existing:
		large_delete_all( SeasonsPassHolder, Q(seasons_pass=seasons_pass) )
	
	ifm = standard_field_map()
	
	# Process the records in large transactions for efficiency.
	@transaction.atomic
	def process_ur_records( ur_records ):
		for i, ur in ur_records:
		
			v = ifm.finder( ur )
			date_of_birth	= v('date_of_birth', None)
			try:
				date_of_birth = date_from_value(date_of_birth)
			except Exception as e:
				ms_write( f'Row {i:>6}: Ignoring birthdate (must be YYYY-MM-DD) "{date_of_birth}" ({ur}) {e}' )
				date_of_birth = None
			
			date_of_birth 	= date_of_birth if date_of_birth != invalid_date_of_birth else None
			license_code	= to_int_str(v('license_code', '')).upper().strip()
			last_name		= to_str(v('last_name',''))
			first_name		= to_str(v('first_name',''))
			uci_id			= to_uci_id(v('uci_id', None))

			license_holder = None
			if not license_holder and license_code:
				license_holder = LicenseHolder.objects.filter( license_code=license_code ).first()
			else:
				if not license_holder and uci_id:
					license_holder = LicenseHolder.objects.filter( uci_id=uci_id ).first()
					
				if not license_holder and last_name:
					q = Q(search_text__startswith=utils.get_search_text([last_name,first_name]))
					if date_of_birth:
						q &= Q(date_of_birth=date_of_birth)				
					license_holder = LicenseHolder.objects.filter( q ).first()

			if not license_holder:
				ms_write( f'Row {i:>6}: Cannot find License Holder by License Code or LastName, FirstName [Date of Birth]\n' )
				continue
				
			seasons_pass.add( license_holder )
			ms_write(
				'Row {i:>6}: {license_code:>8} {dob:>10} {last_name}, {first_name}, {city}, {state_prov}\n'.format(
					i=i,
					license_code=license_holder.license_code,
					dob=license_holder.date_of_birth.strftime('%Y-%m-%d'),
					last_name=license_holder.last_name,
					first_name=license_holder.first_name,
					city=license_holder.city, state_prov=license_holder.state_prov,
				)
			)
	
	sheet_name = None
	if worksheet_contents is not None:
		wb = load_workbook( filename = BytesIO(worksheet_contents), read_only=True, data_only=True )
	else:
		try:
			fname, sheet_name = worksheet_name.split('$')
		except Exception:
			fname = worksheet_name
		wb = load_workbook( filename = fname, read_only=True, data_only=True )
	
	try:
		sheet_name = sheet_name or wb.sheetnames[0]
		ws = wb[sheet_name]
		ms_write( 'Reading sheet "{}"\n'.format(sheet_name) )
	except Exception:
		ms_write( 'Cannot find sheet "{}"\n'.format(sheet_name) )
		return
		
	ur_records = []
	for r, row in enumerate(ws.iter_rows()):
		if r == 0:
			# Get the header fields from the first row.
			fields = ['{}'.format(f.value).strip() for f in row]
			ifm.set_headers( fields )
			ms_write( 'Header Row:\n' )
			for col, f in enumerate(fields, 1):
				name = ifm.get_name_from_alias( f )
				if name is not None:
					ms_write( '        {}. {} --> {}\n'.format(col, f, name) )
				else:
					ms_write( '        {}. ****{} (Ignored)\n'.format(col, f) )
			continue
			
		ur_records.append( (r+1, [v.value for v in row]) )
		if len(ur_records) == 300:
			process_ur_records( ur_records )
			ur_records = []
			
	process_ur_records( ur_records )
	
	ms_write( '\n' )
	ms_write( 'Initialization in: {}\n'.format(datetime.datetime.now() - tstart) )
