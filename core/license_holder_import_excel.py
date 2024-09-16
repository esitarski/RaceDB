import io
import re
import sys
import html
import datetime
import operator
from collections import namedtuple, defaultdict
from openpyxl import load_workbook

from django.db import transaction, IntegrityError
from django.db.models import Q
from django.forms.models import model_to_dict

from . import import_utils
from .import_utils import *
from .CountryIOC import ioc_from_country
from .FieldMap import standard_field_map, normalize
from .get_id import get_id
from .models import *

class tag:
	def __init__( self, stream, name, attr=None ):
		self.stream = stream
		self.name = name
		self.attr = attr
	def __enter__(self):
		self.stream.write( '<{}'.format(self.name) )
		if self.attr:
			self.stream.write( ' {}'.format(' '.join( '{}="{}"'.format(k,v) for k,v in self.attr.items()) ) )
		self.stream.write( '>' )
		return self
	def __exit__(self, type, value, traceback):
		self.stream.write( '</{}>'.format(self.name) )

def license_holder_msg_to_html( msg ):
	s = StringIO()
	icon = {
		1:	'<span class="is-warn">',
		2:	'<span class="is-err">',
	}
	bg_class = {
		1:	{'class':'bg-warning'},
		2:	{'class':'bg-danger'},
	}
	with tag(s, 'table', {
			'class':'table table-hover table-sm table-condensed',
			'style':'border-width: 0;',
		}):
		with tag(s, 'thead'):
			with tag(s, 'tr'):
				with tag(s, 'th'):
					s.write( '' )
				with tag(s, 'th'):
					s.write( 'Row' )
				with tag(s, 'th'):
					s.write( 'Message' )
		
		def write_row( m, bg, icon_str ):
			with tag(s, 'tr', bg):
				with tag(s, 'td'):
					s.write( icon_str )
				row = ''
				if m.startswith('Row'):
					match = re.search(r'^Row\s+(\d+):', m)
					row = int(match.group(1))
					m = m[len(match.group(0)):].strip()
				with tag(s, 'td', {'class':'text-right'}):
					s.write( '{}'.format(row) )
				with tag(s, 'td'):
					m = html.escape( m )
					m = re.sub(r'Added:|Unchanged:|Changed:|Updated:|Warning:|Error:',
						lambda match: '<strong>{}</strong>'.format(match.group(0)), m, flags=re.I)
					s.write( m )
		
		# Write out all errors and warnings first.
		for m in msg.split('\n'):
			m = m.strip()
			icon_str = ''
			bg = None
			for type in range(2, 0, -1):
				if m.startswith('*'*type):
					m = m[type:].strip()
					icon_str = icon.get(type, '')
					bg = bg_class.get(type, None)
					write_row( m, bg, icon_str )
					break
		
		# Then write out all informatoin.
		for m in msg.split('\n'):
			m = m.strip()
			icon_str = ''
			bg = None
			for type in range(2, 0, -1):
				if m.startswith('*'*type):
					m = m[type:].strip()
					icon_str = icon.get(type, '')
					bg = bg_class.get(type, None)
					break
			if bg is None:
				write_row( m, bg, icon_str )
	
	return s.getvalue()

def license_holder_import_excel(
		worksheet_name='', worksheet_contents=None, message_stream=sys.stdout,
		update_license_codes=False,
		set_team_all_disciplines=False,
		replace_bibs=False,
		replace_tags=False,
	):
	system_info = SystemInfo.get_singleton()
	tstart = datetime.datetime.now()
	
	team_lookup = TeamLookup()
		
	license_holder_team = {}
	disciplines = list(Discipline.objects.all())
	effective_date = timezone.localtime(timezone.now()).date()
	def process_license_holder_team( license_holder_team ):
		if license_holder_team:
			TeamHint.objects.filter( license_holder__in=list(license_holder_team.keys()) ).delete()
			team_hints = []
			for lh, t in license_holder_team.items():
				for d in disciplines:
					team_hints.append( TeamHint(license_holder=lh, team=t, discipline=d, effective_date=effective_date) )
			TeamHint.objects.bulk_create( team_hints )
			license_holder_team.clear()
		
	license_holder_discipline_team = {}		# Key: (license_holder, discipline), Data: team.
	def process_license_holder_discipline_team( license_holder_discipline_team ):
		for (license_holder, discipline), team in license_holder_discipline_team.items():
			team_hint = TeamHint.objects.filter(license_holder=license_holder, discipline=discipline).order_by('-effective_date').first()
			if team_hint:
				TeamHint.objects.filter(license_holder=license_holder, discipline=discipline).exclude(id=team_hint.id).delete()
				team_hint.team = team
				team_hint.effective_date = effective_date
			else:
				team_hint = TeamHint(license_holder=license_holder, discipline=discipline, effective_date=effective_date, team=team)
			team_hint.save()
		license_holder_discipline_team.clear()

	Info, Warning, Error = 0, 1, 2
	prefix = {i:'*'*i for i in range(3)}
	if message_stream == sys.stdout or message_stream == sys.stderr:
		def ms_write( s, flush=False, type=Info ):
			s = prefix[type] + '{}'.format(s)
			message_stream.write( removeDiacritic(s) )
	else:
		def ms_write( s, flush=False, type=Info ):
			s = prefix[type] + '{}'.format(s)
			message_stream.write( s )
			sys.stdout.write( removeDiacritic(s) )
			if flush:
				sys.stdout.flush()
	
	#-------------------------------------------------------------------------------------------------
	discipline_cols = {
		'Road':				['national road', 'provincial road'],
		'Cyclocross':		['national cyclocross', 'provincial cyclocross'],
		'Track':			['track'],
		'Gravel':			['gravel'],
		'MTB':				['cross country', 'provincial cross country', 'downhill', 'fourx'],
		'Para':				['para cycling'],
	}
	discipline_by_name = {}
	for dname in discipline_cols.keys():
		try:
			discipline_by_name[dname] = Discipline.objects.get(name=dname)
		except (Discipline.DoesNotExist,  Discipline.MultipleObjectsReturned):
			pass
	#-------------------------------------------------------------------------------------------------
	
	cpy_prefix = '_CPY_'
	dup_prefix = '_DUP_'
	err_prefix = '_XXX_'
	
	# Clean up all existing cpy or dup codes.
	success = True
	while success:
		success = False
		with transaction.atomic():
			for lh in LicenseHolder.objects.filter( license_code__startswith=cpy_prefix )[:999]:
				lh.license_code = lh.license_code.replace( cpy_prefix, err_prefix )
				lh.save()
				success = True
	success = True
	while success:
		success = False
		with transaction.atomic():
			for lh in LicenseHolder.objects.filter( license_code__startswith=dup_prefix )[:999]:
				lh.license_code = lh.license_code.replace( dup_prefix, err_prefix )
				lh.save()
				success = True

	ifm = standard_field_map()
	disciplines = list( Discipline.objects.all() )
	for d in disciplines:
		aliases = ['{} Team'.format(d.name)]
		if d.name == 'Cyclocross':
			aliases.append( 'CX Team' )
		ifm.set_aliases( d.name, aliases )
	discipline_teams = []
	
	status_count = defaultdict( int )
	
	license_code_aliases = []
	
	def ensure_unique( license_code, tag, existing_id=None ):
		if license_code:
			q = LicenseHolder.objects.filter(license_code=license_code)
			if existing_id:
				q = q.exclude( id=existing_id )
			lh = q.first()
			if lh:
				lh.license_code = random_temp_license()
				lh.save()			
		if tag:
			q = LicenseHolder.objects.filter(existing_tag=tag)
			if existing_id:
				q = q.exclude( id=existing_id )
			lh = q.first()
			if lh:
				lh.existing_tag = get_id( system_info.tag_bits )
				lh.save()
		
	# Process the records in large transactions for efficiency.
	today = datetime.date.today()
	
	def v_to_str( k, v ):
		# Return a reasonable string field.
		if isinstance(v, str):
			return f'"{v}"'
		if k == 'gender':
			return ['Men','Women'][v]
		return str(v)
				
	def clean_license_header_row( i, ur ):
		v = ifm.finder( ur )
		
		#-----------------------------------------------------------------------------------------
		# Process all the fields from the Excel sheet.
		#
		uci_code		= to_str(v('uci_code', None))
		date_of_birth	= v('date_of_birth', None)
		age             = to_int(v('age', None))
		
		try:
			date_of_birth = date_from_value(date_of_birth)
		except Exception as e:
			ms_write( 'Row {}: Warning: Ignoring birthdate (must be YYYY-MM-DD) "{}" ({}) {}'.format(
				i, date_of_birth, ur, e), type=Warning,
			)
			date_of_birth = None
		
		# If no date of birth, try to get it from the UCI code.
		if not date_of_birth and uci_code:
			try:
				date_of_birth = datetime.date( int(uci_code[3:7]), int(uci_code[7:9]), int(uci_code[9:11]) )
			except Exception:
				pass
		
		# If no date of birth, make one up based on the age.
		if not date_of_birth and age:
			date_of_birth = date_of_birth_from_age( age, today )
			year_only_dob = True
		else:
			year_only_dob = False
		
		# As a last resort, pick the default DOB
		date_of_birth 	= date_of_birth or invalid_date_of_birth
		
		license_code = None
		for lc in license_code_aliases:
			license_code = to_int_str(v(lc, None))
			if license_code is not None:
				license_code = license_code.upper().strip().split(',')[0]
				if license_code:
					break
		if not license_code or license_code == 'TEMP':
			license_code = None
		
		last_name		= to_str(v('last_name',''))
		first_name		= to_str(v('first_name',''))
		
		gender			= to_str(v('gender',''))
		gender			= gender_from_str(gender) if gender else None
		
		email			= to_str(v('email', None))
		phone			= to_phone(v('phone', None))
		city			= to_str(v('city', None))
		state_prov		= to_str(v('state_prov', None))
		zip_postal		= to_str(v('zip_postal', None))
		if zip_postal:
			zip_postal = validate_postal_code( zip_postal )
		nationality		= to_str(v('nationality', None))
		
		uci_id			= to_uci_id(v('uci_id', None))
		nation_code		= to_str(v('nation_code', None))
		
		bib				= (to_int(v('bib', None)) or None)
		existing_bib	= (to_int(v('existing_bib', None)) or None)
		
		tag				= to_int_str(v('tag', None))
		existing_tag	= to_int_str(v('existing_tag', None))
		existing_tag2	= to_int_str(v('existing_tag2', None))
		
		note		 	= to_str(v('note', None))
		
		emergency_contact_name = to_str(v('emergency_contact_name', None))
		emergency_contact_phone = to_phone(v('emergency_contact_phone', None))
		emergency_medical = to_str(v('emergency_medical', None))

		team_name		= to_str(v('team', None))
		club_name		= to_str(v('club', None))
		if not team_name:
			team_name = club_name
		team_code		= to_str(v('team_code', None))
		
		# Override the exisiting fields if we have specific data.
		if tag and not existing_tag:
			existing_tag = tag
		if bib and not existing_bib:
			existing_bib = bib

		#---------------------------------------------
		
		license_holder_attr_value = {
			'license_code':license_code,
			'last_name':last_name,
			'first_name':first_name,
			'gender':gender,
			'date_of_birth':date_of_birth,
			'year_only_dob':year_only_dob,
			
			'uci_code':uci_code,
			'email':email,
			'phone':phone,
			'city':city,
			'state_prov':state_prov,
			'nationality':nationality,
			'zip_postal':zip_postal,
			
			'nation_code':nation_code,
			'uci_id':uci_id,
			
			'existing_tag':existing_tag if replace_tags else None,
			'existing_tag2':existing_tag2 if replace_tags else None,
			
			'existing_bib':existing_bib if replace_bibs else None,
			
			'note':note,
			
			'emergency_contact_name':emergency_contact_name,
			'emergency_contact_phone':emergency_contact_phone,
			'emergency_medical':emergency_medical,

			'team_name':team_name,
			'team_code':team_code,
		}
		return i, { a:v for a, v in license_holder_attr_value.items() if v }
	
	def process_license_header_rows( license_holder_rows ):
		# Put all the cleansed data into the database in one transaction.
		# It is critical not to throw any exceptions here.
		
		with transaction.atomic():
			for i, lhr in license_holder_rows:			
				#------------------------------------------------------------------------------
				# Update Team
				#
				team = None
				team_name = lhr.pop('team_name', None)
				team_code = lhr.pop('team_code', None)
				
				team_args = { 'name':team_name, 'team_code':team_code }
				team_args = {k:v for k,v in team_args.items() if v}
				
				if team_name not in team_lookup:
					msg = 'Row {:>6}: Added team: {}\n'.format(
						i,
						', '.join( '{}'.format(v) for v in [team_name, team_code,] ),
					)
					ms_write( msg )
				team = team_lookup[team_name]
				if team and set_attributes_changed( team, team_args, False ):
					msg = 'Row {:>6}: Updated team: {}\n'.format(
						i,
						', '.join( '{}'.format(v) for v in [team_name, team_code,] ),
					)
					ms_write( msg )
					truncate_char_fields(team).save()

				# Get the teams for each discipline.
				d_team = []
				for d in discipline_teams:
					tn = lhr.pop( d.name, None )
					if not tn:							# Skip empty field.
						continue
					if Team.is_independent_name(tn):	# Explicitly handle independent as null team.
						d_team.append( (d, None) )
						continue
					if tn not in team_lookup:
						msg = 'Row {:>6}: Added team: {}\n'.format(
							i,
							', '.join( '{}'.format(v) for v in [team_name, team_code,] ),
						)
						ms_write( msg )
					d_team.append( (d, team_lookup[tn]) )
				
				#------------------------------------------------------------------------------
				# Get LicenseHolder.
				#
				license_holder = None
				status = 'Unchanged'
				
				year_only_dob = lhr.pop('year_only_dob', False)
				
				last_name = lhr.get('last_name','')
				first_name = lhr.get('first_name','')
				name = ' '.join( n for n in (first_name, last_name) if n )
				
				gender = lhr.get('gender', None)
				date_of_birth = lhr.get('date_of_birth', None)
				
				license_code = lhr.get('license_code', None)
				uci_id = lhr.get('uci_id', None)
				existing_tag = lhr.get('existing_tag', None)
				
				# Get a query by Last, First, DOB and Gender.
				qNameDOBGender = Q( search_text__startswith=utils.get_search_text([last_name, first_name]) )
				if date_of_birth and date_of_birth != invalid_date_of_birth:
					if year_only_dob:
						qNameDOBGender &= Q( date_of_birth__year=date_of_birth.year )
					else:
						qNameDOBGender &= Q( date_of_birth=date_of_birth )
				if gender is not None:
					qNameDOBGender &= Q( gender=gender )
				
				#------------------------------------------------------------------------------
				if not license_holder and uci_id:
					lhs = list( LicenseHolder.objects.filter(uci_id=uci_id) )
					if len(lhs) == 1:
						license_holder = lhs[0]
					elif len(lhs) > 1:
						ms_write( 'Row {}: Warning:  Name="{}" found duplicate UCI ID="{}"\n'.format(
								i, name, uci_id,
							), type=Warning
						)
				
				#------------------------------------------------------------------------------
				if not license_holder and update_license_codes:
					# Try to find the license holder by name, DOB, gender
					lhs = list( LicenseHolder.objects.filter(qNameDOBGender) )
					if len(lhs) == 1:
						license_holder = lhs[0]
						ensure_unique( license_code, existing_tag, license_holder.id )
					
					elif len(lhs) == 0:
						# Create a new license holder from the information given.
						if not license_code:
							lhr['license_code'] = license_code = random_temp_license()
						
						ensure_unique( license_code, existing_tag )
						license_holder = LicenseHolder( **lhr )
						truncate_char_fields(license_holder).save()
						status = 'Added'
					
					else:
						ms_write( 'Row {}: Warning:  Update not performed.  Found multiple LicenceHolders matching "Last, First DOB Gender" Name="{}"\n'.format(
								i, name,
							), type=Warning
						)
						continue
						
					
				#------------------------------------------------------------------------------
				if not license_holder and license_code:
					# Try to find the license holder by license code.
					lhs = list( LicenseHolder.objects.filter(license_code=license_code) )
					if len(lhs) == 1:
						license_holder = lhs[0]
					elif len(lhs) == 0:
						# No match.  Create a new license holder using the given license code.
						ensure_unique( license_code, existing_tag )
						license_holder = LicenseHolder( **lhr )
						truncate_char_fields(license_holder).save()
						del lhr['license_code'] # Delete license_code after creating a record to prevent update.
						status = 'Added'
							
				if not license_holder:
					# Try to find the license holder by name, DOB, gender
					lhs = list( LicenseHolder.objects.filter(qNameDOBGender) )
					if len(lhs) == 1:
						license_holder = lhs[0]
					elif len(lhs) == 0:
						# No name match.
						# Create a temporary license holder.
						lhr['license_code'] = license_code = random_temp_license()
						ensure_unique( license_code, existing_tag )
						license_holder = LicenseHolder( **lhr )
						truncate_char_fields(license_holder).save()
						del lhr['license_code'] # Delete license_code after creating a record to prevent the update.
						status = 'Added'
					
					elif len(lhs) > 1:
						ms_write( 'Row {}: Error:  found multiple LicenceHolders matching "Last, First, DOB, Gender" Name="{}"\n'.format(
								i, name,
							), type=Error,
						)
						continue
					
				if not license_holder:
					ms_write( 'Row {}: Error:  Cannot find License Holder: Name="{}"\n'.format(
							i, name,
						), type=Error,
					)
					continue
				
				fields_changed = set_attributes_changed( license_holder, lhr, False )
				if fields_changed:
					ensure_unique( license_code, existing_tag, license_holder.id )
					truncate_char_fields(license_holder).save()
					if status != 'Added':
						status = 'Changed'
				
				status_count[status] += 1
				if status != 'Unchanged':
					msg = 'Row {:>6}: {}: {:>8}: {}\n'.format(
						i,
						status,
						license_holder.license_code,
						', '.join( f'{k}={v_to_str(k,v)}' for k,v in model_to_dict( license_holder ).items() if k not in ('id','search_text') and v not in (None,'') )
					)
					if status == 'Changed':
						msg += '            Updated: {}\n'.format( ', '.join( '{}=({})'.format(k,v) for k,v in fields_changed ) )
					ms_write( msg )
					
				if license_holder and d_team:
					for discipline, td in d_team:
						license_holder_discipline_team[(license_holder, discipline)] = td
					if len(license_holder_discipline_team) >= 200:
						process_license_holder_discipline_team( license_holder_discipline_team )
						
				if license_holder and team_name and set_team_all_disciplines:
					license_holder_team[license_holder] = team
					if len(license_holder_team) >= 200:
						process_license_holder_team( license_holder_team )

	sheet_name = None
	if worksheet_contents is not None:
		wb = load_workbook( filename = io.BytesIO(worksheet_contents), read_only=True, data_only=True )
	else:
		try:
			fname, sheet_name = worksheet_name.split('$')
		except Exception:
			fname = worksheet_name
		wb = open_workbook( filename = fname, read_only=True, data_only=True )
	
	try:
		sheet_name = sheet_name or wb.sheetnames[0]
		ws = wb[sheet_name]
		ms_write( 'Reading sheet "{}"\n'.format(sheet_name) )
	except Exception:
		ms_write( 'Cannot find sheet "{}"\n'.format(sheet_name) )
		return
	
	license_holder_rows = []
	
	for r, row in enumerate(ws.iter_rows()):
		if r == 0:
			# Get the header fields from the first row.
			fields = ['{}'.format(f.value).strip() for f in row]
			ifm.set_headers( fields )
			license_code_aliases = [lh for lh in ifm.name_to_col.keys() if lh.startswith('license_code')]
			discipline_teams = [d for d in disciplines if d.name in ifm]
			
			ms_write( 'Header Row:\n' )
			for col, f in enumerate(fields, 1):
				name = ifm.get_name_from_alias( f )
				if name is not None:
					ms_write( '        {}. {} --> {}\n'.format(col, f, name) )
				else:
					ms_write( '        {}. ****{} (Ignored)\n'.format(col, f) )
			ms_write( '\n' )
			continue
			
		license_holder_rows.append( clean_license_header_row(r+1, [v.value for v in row]) )
		if len(license_holder_rows) == 300:
			process_license_header_rows( license_holder_rows )
			license_holder_rows[:] = []
			
	process_license_header_rows( license_holder_rows )
	
	process_license_holder_team( license_holder_team )
	process_license_holder_discipline_team( license_holder_discipline_team )
		
	ms_write( '\n' )
	ms_write( '   '.join( '{}: {}'.format(a, v) for a, v in sorted((status_count.items()), key=operator.itemgetter(0)) ) )
	ms_write( '\n' )
	ms_write( 'Initialization in: {}\n'.format(datetime.datetime.now() - tstart) )
	ms_write( '', True )
